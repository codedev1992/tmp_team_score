import streamlit as st 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from io import BytesIO
from openpyxl.styles import Alignment, Font
from openpyxl.utils import range_boundaries
from openpyxl.styles import PatternFill
import math
import random

DEBUG = True

teams_file = st.file_uploader("upload teams excel file.")

st.info("NOTE: For generating My Teams, the player value should start from A19")
c1,c2 = st.columns(2)
with c1:
    process_button = st.button("process")
with c2:
    my_team_formation = st.button("My Team Process")

players_sheet_name  = "PlayersList"
sheet_to_use = "Copy of Teams"
teams_list_sheet_name = "TeamsList"
my_team_sheet_name = "My Teams"

C_COLOR = ["FF00FF00","FF00B050"]
VC_COLOR = ["FFFFFF00"] 

TEAMS = []


PLAYER_TYPE_RULES = {
"W": {"min" : 1, "max": 4},
"Ba" :{"min" : 3, "max": 6},
"A": {"min" : 1, "max": 4},
"Bo": {"min" : 3, "max": 6},
}

def plyer_to_type(team,mapping):
    team_type_count = {"W":0, "Ba":0, "Bo": 0, "A":0}

    for pl in team:
        team_type_count[mapping[pl].get("type")] = team_type_count[mapping[pl].get("type")] + 1
    
    return team_type_count

def get_team_combination(no_team,player_to_type):

    output = {
        "W" :[],
        "Ba":[],
        "Bo":[],
        "A":[]
    }

    total_possible_players = no_team * 12 
    w,ba,bo,a = 2,4,4,2

    w1,ba1,bo1,a1 = len(player_to_type["W"]),len(player_to_type["Ba"]),len(player_to_type["Bo"]),len(player_to_type["A"])

    exp_w, exp_ba, exp_bo, exp_a = w*no_team, ba*no_team, bo*no_team, a*no_team

    wgt_w, wgt_ba,wgt_bo,wgt_a = math.ceil(exp_w/w1),math.ceil(exp_ba/ba1),math.ceil(exp_bo/bo1),math.ceil(exp_a/a1)

    diff_weight = {
        "W":(wgt_w*w1)-exp_w, 
        "Ba":(wgt_ba*ba1)-exp_ba,
        "Bo":(wgt_bo*bo1)-exp_bo,
        "A":(wgt_a*a1)-exp_a
    }

    p_weights = {
        "W":[wgt_w]*w1,
        "Ba":[wgt_ba]*ba1,
        "Bo":[wgt_bo]*bo1,
        "A":[wgt_a]*a1
    }

     
    for k,v in diff_weight.items():
        v = diff_weight[k]
        if v > 0:
            for knam,vidx in diff_weight.items():
                new_va= p_weights[knam][vidx] - 1
               
                p_weights[knam][vidx]  = new_va

    p_weights_to_dict = p_weights

    for k,v in p_weights_to_dict.items():
        for idx in range(len(v)):
            output[k].append([player_to_type[k][idx]]*v[idx])

    return output

def generate_my_teams(exel_file):
    no_team = st.session_state["input_team_generation_count"]

    RED = "FFFF0000"
    BLACK = 'FF000000'
    wb = load_workbook(exel_file, read_only=False)
    sheet_names = wb.sheetnames
    if my_team_sheet_name in sheet_names:
        my_team_players = {}
        my_team_sheet = wb[my_team_sheet_name]
        gph_idx = f"A19:B40"
        
        rnk = 1 
        top_11_players = []
        bottom_11_player = []
        player_to_type = {
            "A":[],
            "Ba" :[],
            "Bo":[],
            "W":[]
        }
        for row in my_team_sheet[gph_idx]:
            if row[0] is not None:
                
                if row[0].font and row[0].font.color:
                    font_color = row[0].font.color.rgb
                    if font_color  == RED:
                        color = "r"
                    else:
                        color = "b"
                else:
                    color = "b"
                
                my_team_players[row[0].value] = {"color": color, "rank" :rnk, "type":  row[1].value}

                player_to_type[row[1].value].append(row[0].value)

                if rnk <=11:
                    top_11_players.append(row[0].value)
                else:
                    bottom_11_player.append(row[0].value)
                rnk = rnk + 1

        team_comb_dict = get_team_combination(no_team,player_to_type)

        play_expt_cmb_cnt = {"min":{"W":1,"Ba":3,"Bo":3,"A":1},"max":{"W":2,"Ba":4,"Bo":4,"A":2}}

        for k, v in team_comb_dict.items():
            print(k,len(v))

        my_team = []

        for i in range(no_team):
            team = []
            #print(team_comb_dict)
            for k,cnt in play_expt_cmb_cnt["min"].items():
                for j in range(cnt):
                    #print(i,k,team_comb_dict[k])
                    pname = team_comb_dict[k][j].pop(0)
                    # for pidx in range(len(team_comb_dict[k])):
                    #     if len(team_comb_dict[k][pidx]) > 0:
                            
                    #         break
                    team.append(pname)
            #random.shuffle(team)
            my_team.append(team)
            #random.shuffle(my_team)
             

            for player, types in team_comb_dict.items():
                team_comb_dict[player] = [lst for lst in types if lst]

            # for k, v in team_comb_dict.items():
            #     idx_to_pop = []
            #     for idx in range(len(team_comb_dict[k])):
            #         if len(team_comb_dict[k][idx]) == 0 :
            #             print("popped", k, idx)
            #             idx_to_pop.append(idx)
                
            #     for idx in idx_to_pop:
            #         team_comb_dict[k].pop(idx)

        for i in range(no_team):
            for p_type in ["Ba","Bo","A","W"]:
                p_type_cnt = plyer_to_type(my_team[i], my_team_players)
                for plist in team_comb_dict[p_type]:
                    if (len(my_team[i]) <11 and 
                        len(team_comb_dict[p_type]) > 0 and 
                        p_type_cnt[p_type] < play_expt_cmb_cnt["max"].get(p_type)):

                        if (plist and plist[0] not in my_team[i]):
                            my_team[i].append(plist[0])
                            plist.pop(0)

                for player, types in team_comb_dict.items():
                    team_comb_dict[player] = [lst for lst in types if lst]
        
        
        for t in my_team:
            random.shuffle(t)
        
        random.shuffle(my_team)
          
        team_count = 1
        last_col_name = get_column_letter(no_team)
        write_range = f"A1:{last_col_name}11"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
                    
        teams_status = []
        t_count = 1
        #print(my_team_players)
        for update_team in my_team:
            tems_cnt = {"W":0, "Ba": 0, "A": 0, "Bo": 0,"r":0,"b":0}
            for pname in update_team:
                pyr = my_team_players.get(pname,{})
                ptype = pyr.get("type","W")
                pcolor = pyr.get("color","b")
                if ptype is not None:
                    tems_cnt[ptype] = tems_cnt[ptype] + 1
                
                if pcolor is not None:
                    tems_cnt[pcolor] = tems_cnt[pcolor] + 1
        
            teams_status.append(tems_cnt)
            update_team.insert(0,t_count)
            t_count = t_count + 1
                
        # print(min_col, min_row, max_col, max_row,len(my_team), len(my_team[0]))
        for col in range(min_col, max_col + 1):
            for row in range(min_row, max_row + 1):
                pname = my_team[col-1][row-1]
                cell = my_team_sheet.cell(row=row, column=col)
                
                if pname in bottom_11_player:
                    cell.fill = PatternFill(start_color="FBDAD7", fill_type='solid')
                f_color = my_team_players.get(pname,{}).get("color","b")
                if f_color == "r":
                    red_font = Font(color=RED) 
                    cell.font = red_font
                else:
                    black_font = Font(color=BLACK) 
                    cell.font = black_font

                cell.value = pname
        
        write_range = f"A13:{last_col_name}18"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        for col in range(min_col, max_col + 1):
            tm_status = teams_status[col-1]

            a_count = tm_status.get("A", 0) 
            w_count =  tm_status.get("W", 0) 
            ba_count =   tm_status.get("Ba", 0) 
            bo_count  =  tm_status.get("Bo", 0)

            r_colr_count = tm_status.get("r", 0)
            b_colr_count = tm_status.get("b", 0)

            a_count_str = "A "+ str(a_count)
            w_count_str = "W "+ str(w_count)
            ba_count_str = "Ba "+ str(ba_count)
            bo_count_str = "Bo "+ str(bo_count)

            r_colr_count_str = "Red "+ str(r_colr_count)
            b_colr_count_str = "Black "+ str(b_colr_count)

            cell = my_team_sheet.cell(row=13, column=col)
            cell.value = a_count_str + "," + w_count_str

            cell = my_team_sheet.cell(row=14, column=col)
            cell.value = ba_count_str + "," + bo_count_str

            not_perfects = []

            if not (PLAYER_TYPE_RULES["A"]["min"] <= a_count  and a_count <= PLAYER_TYPE_RULES["A"]["max"]):
                not_perfects.append(a_count_str) 
            
            if not (PLAYER_TYPE_RULES["W"]["min"] <= w_count  and w_count <= PLAYER_TYPE_RULES["W"]["max"]):
                not_perfects.append(w_count_str) 
            
            if not (PLAYER_TYPE_RULES["Ba"]["min"] <= ba_count  and ba_count <= PLAYER_TYPE_RULES["Ba"]["max"]):
                not_perfects.append(ba_count_str) 
            
            if not (PLAYER_TYPE_RULES["Bo"]["min"] <= bo_count  and bo_count <= PLAYER_TYPE_RULES["Bo"]["max"]):
                not_perfects.append(bo_count_str) 
                
            if len(not_perfects) == 0:
                cell = my_team_sheet.cell(row=15, column=col)
                cell.value = "Perfect"
                black_font = Font(color=BLACK) 
                cell.font = black_font
            else:
                cell = my_team_sheet.cell(row=15, column=col)
                red_font = Font(color=RED) 
                cell.font = red_font
                cell.value = "Not Perfect"

                cell = my_team_sheet.cell(row=16, column=col)
                cell.value = ",".join(not_perfects)
            
            cell = my_team_sheet.cell(row=17, column=col)
            cell.value = r_colr_count_str + "," + b_colr_count_str


        team_output = BytesIO()
        wb.save(team_output)
        team_output.seek(0)

        file_name = exel_file.name
        # Step 4: Create a download button
        btn = st.download_button(
            label="Download Excel with My Team Formation",
            data=team_output,
            file_name="my_team_updated_"+file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

####Depricated
        # my_team = []
        # print(no_team)
        # for team_id in range(no_team):
        #     my_team.append(list(top_11_players))
        

        # pvt_point = int(no_team / 2)

        # s_idx, m_idx, l_idx = 0, pvt_point, no_team
        # p_idx, itr  = 0, 1
        # rest_player_idx_at = no_team if no_team > len(bottom_11_player) else len(bottom_11_player) 
        # for i in range(10,-1, -1):

        #     if i % 2 == 0:
        #         s_idx, m_idx = 0, pvt_point
        #     else:
        #         s_idx = m_idx
        #         m_idx = l_idx
        
        #     for j in range(s_idx,m_idx):
                
        #         my_team[j][i] = bottom_11_player[p_idx%11]
        #         p_idx = p_idx + 1
            
        #         if itr % rest_player_idx_at == 0:
        #             bottom_11_player = bottom_11_player[1:] + [bottom_11_player[0]]

        #         itr = itr + 1

        # team_count = 1
        # last_col_name = get_column_letter(no_team)
        # write_range = f"A1:{last_col_name}12"
        # min_col, min_row, max_col, max_row = range_boundaries(write_range)
        

        # teams_status = []
        # t_count = 1
        # #print(my_team_players)
        # for update_team in my_team:
        #     tems_cnt = {"W":0, "Ba": 0, "A": 0, "Bo": 0,"r":0,"b":0}
        #     for pname in update_team:
        #         pyr = my_team_players.get(pname,{})
        #         ptype = pyr.get("type","W")
        #         pcolor = pyr.get("color","b")
        #         if ptype is not None:
        #             tems_cnt[ptype] = tems_cnt[ptype] + 1
                
        #         if pcolor is not None:
        #             tems_cnt[pcolor] = tems_cnt[pcolor] + 1
        
        #     teams_status.append(tems_cnt)
        #     update_team.insert(0,t_count)
        #     t_count = t_count + 1
                
        # # print(min_col, min_row, max_col, max_row,len(my_team), len(my_team[0]))
        # for col in range(min_col, max_col + 1):
        #     for row in range(min_row, max_row + 1):
        #         pname = my_team[col-1][row-1]
        #         cell = my_team_sheet.cell(row=row, column=col)
                
        #         if pname in bottom_11_player:
        #             cell.fill = PatternFill(start_color="FBDAD7", fill_type='solid')
        #         f_color = my_team_players.get(pname,{}).get("color","b")
        #         if f_color == "r":
        #             red_font = Font(color=RED) 
        #             cell.font = red_font
        #         else:
        #             black_font = Font(color=BLACK) 
        #             cell.font = black_font

        #         cell.value = pname
            
        # write_range = f"A13:{last_col_name}18"
        # min_col, min_row, max_col, max_row = range_boundaries(write_range)
        # for col in range(min_col, max_col + 1):
        #     tm_status = teams_status[col-1]

        #     a_count = tm_status.get("A", 0) 
        #     w_count =  tm_status.get("W", 0) 
        #     ba_count =   tm_status.get("Ba", 0) 
        #     bo_count  =  tm_status.get("Bo", 0)

        #     r_colr_count = tm_status.get("r", 0)
        #     b_colr_count = tm_status.get("b", 0)

        #     a_count_str = "A "+ str(a_count)
        #     w_count_str = "W "+ str(w_count)
        #     ba_count_str = "Ba "+ str(ba_count)
        #     bo_count_str = "Bo "+ str(bo_count)

        #     r_colr_count_str = "Red "+ str(r_colr_count)
        #     b_colr_count_str = "Black "+ str(b_colr_count)

        #     cell = my_team_sheet.cell(row=13, column=col)
        #     cell.value = a_count_str + "," + w_count_str

        #     cell = my_team_sheet.cell(row=14, column=col)
        #     cell.value = ba_count_str + "," + bo_count_str

        #     not_perfects = []

        #     if not (PLAYER_TYPE_RULES["A"]["min"] <= a_count  and a_count <= PLAYER_TYPE_RULES["A"]["max"]):
        #         not_perfects.append(a_count_str) 
            
        #     if not (PLAYER_TYPE_RULES["W"]["min"] <= w_count  and w_count <= PLAYER_TYPE_RULES["W"]["max"]):
        #         not_perfects.append(w_count_str) 
            
        #     if not (PLAYER_TYPE_RULES["Ba"]["min"] <= ba_count  and ba_count <= PLAYER_TYPE_RULES["Ba"]["max"]):
        #         not_perfects.append(ba_count_str) 
            
        #     if not (PLAYER_TYPE_RULES["Bo"]["min"] <= bo_count  and bo_count <= PLAYER_TYPE_RULES["Bo"]["max"]):
        #         not_perfects.append(bo_count_str) 
                
        #     if len(not_perfects) == 0:
        #         cell = my_team_sheet.cell(row=15, column=col)
        #         cell.value = "Perfect"
        #         black_font = Font(color=BLACK) 
        #         cell.font = black_font
        #     else:
        #         cell = my_team_sheet.cell(row=15, column=col)
        #         red_font = Font(color=RED) 
        #         cell.font = red_font
        #         cell.value = "Not Perfect"

        #         cell = my_team_sheet.cell(row=16, column=col)
        #         cell.value = ",".join(not_perfects)
            
        #     cell = my_team_sheet.cell(row=17, column=col)
        #     cell.value = r_colr_count_str + "," + b_colr_count_str


        # team_output = BytesIO()
        # wb.save(team_output)
        # team_output.seek(0)

        # file_name = exel_file.name
        # # Step 4: Create a download button
        # btn = st.download_button(
        #     label="Download Excel with My Team Formation",
        #     data=team_output,
        #     file_name="my_team_updated_"+file_name,
        #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # )
#### OLD depricated END
                
    else:
        st.warning("My Team input Sheet is missing. Please add and re-run")

def player_credit_from_excel_sheet(sheet):
    players_credits = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
        name, credit = row
        players_credits[name] = credit if credit is not None else 0
    return players_credits

def check_all_team_marked_c_and_vc(sheet):
    last_column_with_data = 0 
    for column in sheet.iter_rows(min_row=1, max_row=2):
        for cell in column:
            if cell.value is not None:
                last_column_with_data = cell.column

    last_col_name = get_column_letter(last_column_with_data)
    
    last_column_with_data = last_column_with_data - 1

    last_row_index = 12
    last_idx = f"{last_col_name}{last_row_index}"

    # # Create the range string
    gph_idx = f"B2:{last_idx}"

    min_col, min_row, max_col, max_row = range_boundaries(gph_idx)
    c_vc_missing = []
    for col in range(min_col, max_col + 1):
        _c, _vc = False, False
        for row in range(min_row, max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value == "E" or cell.value == "" or cell.value is None:
                _c = True
                _vc = True
                break
            else:
                clr = cell.fill.start_color.index
                #print(col, row, clr,C_COLOR,VC_COLOR)
                if clr in C_COLOR:
                    _c = True
                if clr in VC_COLOR:
                    _vc = True
        
        if _c and _vc:
            pass
        else:
            c_vc_missing.append(sheet.cell(row=1, column=col).value)
    
    if len(c_vc_missing) == 0:
        return True, c_vc_missing
    else:
        return False, c_vc_missing

def compute_and_download(excel_data, is_player_sheet_exists):

    player_credit = {}
    for key in st.session_state:
    # Check if the key belongs to your input fields
        if key.startswith("input_"):
            # Access the value of each input field
            value = st.session_state[key]
            # Process the value as needed
            pname = re.sub("input_[\d]+_", "", key)

            if value is None or value == "":
                player_credit[pname] = 0 
            else:
                player_credit[pname] = eval(value)
    
    #st.write(player_credit)

    credits_rows = []
    for r in excel_data:
        crow = []
        for r_val in r:
            (pname, factor), = r_val.items()
            crow.append(player_credit.get(pname,0) * factor)
        credits_rows.append(crow)
    
    
    # Calculate the number of columns
    num_columns = len(credits_rows[0])
    # Initialize a list with zeros for storing column sums
    column_sums = [0] * num_columns

    column_empty = [""] * num_columns

    # Iterate through each row and column to calculate column sums
    for row in credits_rows:
        for i in range(num_columns):
            column_sums[i] += row[i]

    # Append the column sums to the original 2D list
            
    credits_rows.append(column_empty)
    credits_rows.append(column_sums)

    file_name = teams_file.name

    wb = load_workbook(teams_file, read_only=False)
    
    sheet_names = wb.sheetnames
    sheet_name_lower = [s.lower() for s in sheet_names]

    if players_sheet_name.lower() in sheet_name_lower:
        player_sheet = wb[players_sheet_name]
        wb.remove(player_sheet)
        player_sheet = wb.create_sheet(players_sheet_name)
    else:
        player_sheet = wb.create_sheet(players_sheet_name)
    
    #append the new / updated values
    player_sheet.append(["Name", "Credit"])
    for name, credit in player_credit.items():
        player_sheet.append([name, credit])
    
    # Set the column widths for better readability
    for column in player_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        player_sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length
    

    if sheet_to_use in wb.sheetnames:
        sheet = wb[sheet_to_use]

        start_row = 60
        start_rol = 2

        for row_index, row in enumerate(credits_rows, start=start_row):  # Start=1 since Excel rows are 1-indexed
            for col_index, value in enumerate(row, start=start_rol):  # Start=1 for columns as well
                cell = sheet.cell(row=row_index, column=col_index)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)

                cell.value = value
    
    copy_sheet = wb[sheet_to_use]
    dif_in_score_teams = []

    if teams_list_sheet_name.lower() in sheet_name_lower:
        teams_sheet = wb[teams_list_sheet_name]
        idx = 2
        for v in column_sums:
            teams_sheet["B"+str(idx)] = v
            
            if teams_sheet["C"+str(idx)].value and teams_sheet["C"+str(idx)].value != "":
                diff = float(teams_sheet["C"+str(idx)].value) - float(v) 
                #print(teams_sheet["C"+str(idx)].value,v, diff)
                if diff != 0.0 :
                    dif_in_score_teams.append("T"+ str(teams_sheet["A"+str(idx)].value))
            idx = idx + 1

    else:
        teams_sheet = wb.create_sheet(teams_list_sheet_name)
        teams_sheet.append(["TeamName","Computed","Actual"])

        for team_name in TEAMS:
            teams_sheet.append(["T"+str(team_name)])


    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if len(dif_in_score_teams) > 0:
        team_ids = ",".join(dif_in_score_teams)
        st.error("Teams with values mismatch between computed and actual : "+team_ids )
    else:
        st.success("All values are matching (Both Computed and Actual)")

    # Step 4: Create a download button
    btn = st.download_button(
        label="Download updated Excel file",
        data=output,
        file_name="updated_file_"+file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if process_button:
    wb = load_workbook(teams_file, read_only=False)

    is_all_okay, missing_team = check_all_team_marked_c_and_vc(wb[sheet_to_use])
    if is_all_okay:

        sheet_names = wb.sheetnames
        sheet_name_lower = [s.lower() for s in sheet_names]
        
        data = []
        if sheet_to_use in wb.sheetnames:
            sheet = wb[sheet_to_use]
            last_column_with_data = 0 
            for column in sheet.iter_rows(min_row=1, max_row=2):
                for cell in column:
                    if cell.value is not None:
                        last_column_with_data = cell.column

            last_col_name = get_column_letter(last_column_with_data)
            
            last_column_with_data = last_column_with_data - 1

            st.write("team count : ", last_column_with_data)

            last_row_index = 12
            last_idx = f"{last_col_name}{last_row_index}"

            # Create the range string
            gph_idx = f"B2:{last_idx}"

            for row in sheet[gph_idx]:
                r_values = []
                for cell in row:
                    if cell.value is not None:
                        clr = cell.fill.start_color.index
                        if clr in C_COLOR:
                            r_values.append({cell.value:2})
                        elif clr in VC_COLOR:
                            r_values.append({cell.value:1.5})
                        else:
                            r_values.append({cell.value:1})
                    else:
                         r_values.append({"":0})
                data.append(r_values)
            
            teams_name_idx = f"B1:{last_col_name}1"
            #print("teams : ", teams_name_idx)
            for row in sheet[teams_name_idx]:
                for cell in row:
                    TEAMS.append(cell.value)
            


        if players_sheet_name.lower() in sheet_name_lower:
            st.write("Existing Credit Sheet Found.")
            sheet = wb[players_sheet_name]
            player_credit = player_credit_from_excel_sheet(sheet)

            form = st.form("my_form")
            with form:
                st.header("Player Credits")
                i = 1 
                for name in player_credit:
                    input_key = f"input_{i}_{name}"
                    player_credit[name] = st.text_input(f"{i}.{name}", 
                                                                value=player_credit[name],
                                                                key=input_key)
                    i = i + 1
                submit_btn = form.form_submit_button("Submit", on_click=compute_and_download,args=(data,True) )

        else:
            st.write("No Existing Credit Sheet Found.")
            unique_players_dict = {list(x.keys())[0] for l in data for x in l}

            unique_players_list = list(unique_players_dict)

            form = st.form("my_form")

            wb.close()

            with form:
                i = 1
                input_values = {}
                for name in unique_players_list:
                    input_key = f"input_{i}_{name}"
                    
                    if (name != "E" and len(name) > 2):
                        input_values[input_key] = st.text_input(str(i) + "." + name, key=input_key)
                    else:
                        input_values[input_key] = st.text_input(str(i) + "." + name, key=input_key, value= 0)
                    i = i + 1

                submit_btn = form.form_submit_button("Submit", on_click=compute_and_download,args=(data,False) )
        
    else:
        st.write("All Teams Must Have C and VC, Following Teams are having issue. ")
        st.write(missing_team)


            
                
if my_team_formation:
    form = st.form("my_form")
    with form:

        expected_team_count = st.number_input("Enter no of team required.", min_value=3,
                                               max_value=5000, 
                                               step=1,
                                                 format='%d', key="input_team_generation_count")
        submit_btn = form.form_submit_button("Generate Team",
                                              on_click=generate_my_teams,
                                              args=(teams_file,))
        

        
